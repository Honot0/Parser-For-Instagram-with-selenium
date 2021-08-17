from selenium.webdriver import Firefox
from bs4 import BeautifulSoup
import re
from openpyxl import load_workbook
from openpyxl import Workbook



# import urllib.request
# import xml.etree.ElementTree as et
# from itertools import groupby
# import pickle




def Filter_this(lisst):# почему тут так много одинаковых циклов подряд ? а хз シシシシシ. он не хочет просеить мусор в одном почему-от
    lisst = lisst.split("\n")
    pattern4 = r'[а-яА-Я]'
    pattern5 = r','
    pattern6 = r'#'
    pattern7 = r' '
    ind = 0
    for each in lisst:
        if re.search(pattern4, each, flags=re.IGNORECASE) or re.search(pattern5, each, flags=re.IGNORECASE)  or re.search(pattern6, str(each), flags=re.IGNORECASE) or re.search(pattern7, each, flags=re.IGNORECASE):
            lisst.pop(ind)
            print(each)
        ind+=1
    ind = 0
    for each in lisst:
        if re.search(pattern4, each, flags=re.IGNORECASE) or re.search(pattern5, each, flags=re.IGNORECASE)  or re.search(pattern6, str(each), flags=re.IGNORECASE) or re.search(pattern7, each, flags=re.IGNORECASE):
            lisst.pop(ind)
            print(each)
        ind+=1
    ind = 0
    for each in lisst:
        if each.startswith("#"):
            lisst.pop(ind)
            print(each)
        ind+=1
    ind = 0
    for each in lisst:
        if each.startswith("#"):
            lisst.pop(ind)
            print(each)
        ind+=1
    ind = 0
    for each in lisst:
        if each.startswith("#"):
            lisst.pop(ind)
            print(each)
        ind+=1
    ind = 0
    for each in lisst:
        if each==None or each=="" or each=="None"or each=='':
            lisst.pop(ind)
            print(each)
        ind+=1
    ind = 0
    for each in lisst:
        if each==None or each=="" or each=="None"or each=='':
            lisst.pop(ind)
            print(each)
        ind+=1
    return lisst

def make_urls(nicks):
    new_list = []
    for each in nicks:
        new_list.append('https://www.instagram.com/%s/'% (each))
    return new_list



INPUTT = input("Введите поисковый запрос: ")

base_url = 'https://www.instagram.com/directory/profiles/'
browser = Firefox(executable_path="geckodriver", service_log_path="geckodriver.log")
browser.get(base_url)
html = browser.page_source

browser.find_element_by_xpath("/html/body/div[1]/section/nav/div[2]/div/div/div[2]/div/div/span[2]").click()
browser.find_element_by_xpath("/html/body/div[1]/section/nav/div[2]/div/div/div[2]/input").send_keys(INPUTT)

import time
time.sleep(7)

nicks = browser.find_element_by_xpath("/html/body/div[1]/section/nav/div[2]/div/div/div[2]/div[3]/div/div[2]/div")

print(type(nicks.text))
nicks = Filter_this(nicks.text)

urls = make_urls(nicks)
print(urls)




# filename = '%s.txt' % (INPUTT)
# with open(filename, 'w') as f_obj:
#     f_obj.write(".")
#     f_obj.write("\n")

filename = '%s.xlsx' % (INPUTT)
wb = Workbook()
ws = wb.active
ws['A1'] = str(INPUTT)
wb.save(filename)
wb.close()

for url in urls:
    try:
        browser = Firefox(executable_path="geckodriver", service_log_path="geckodriver.log")
        browser.get(url)
        new_html = browser.page_source
        import  time
        time.sleep(1)
        try:
            des_text = browser.find_element_by_xpath("/html/body/div[1]/section/main/div/header/section/div[2]")
            des_text=des_text.text
        except:
            des_text = "no_description"
        pattern1 = r'(\d{1,}(?:\ |\-|\)|\()*){10,}'
        pattern2 = r"[^\n]*(.com|.ws|.ru|.me|api\.|\.store\/|.be|\@)[^\n]*"


        try:
            phone = re.search(pattern1, des_text)[0]
        except:
            phone = "no_phone"

        try:
            site = re.search(pattern2, des_text)[0]
        except:
            site = "no_contact"
        try:
            site2 = re.search(pattern2, des_text)[2]
        except:
            site2 = "no_contact"


        sttt = [str(url), str(phone), str(site),str(site2), str(des_text)]
        print(sttt[0:-1])

        wb = load_workbook(filename=filename, read_only=False)
        ws = wb.active
        ws.append(sttt)
        wb.save(filename)
        wb.close()

        # sttt =' URL: %s PHONE: %s CONTACTS: %s FULL_TEXT:\n %s'%(str(url),str(phone),str(site),str(des_text))
        # my_file = open(filename, "r")
        # content = my_file.read()
        # my_file.close()
        # with open(filename, 'w') as f_obj:
        #     f_obj.write(content)
        #     f_obj.write(sttt)
        #     f_obj.write("\n")
        #     f_obj.write("\n")
    except Exception as e:
        print('Exception in: ',url ," ", e)
        pass
    finally:
        browser.close()








# html = "ффф.txt"
# def get_html(url):
#     if bool(re.search('[а-яА-Я]', url)):
#         response = open(url, encoding='utf-8').read()
#         return response
#     else:
#         response = urllib.request.urlopen(url)
#         return response.read()
# raw = get_html(html)
# print(raw)





# for wq in lisst:
#     if re.search(pattern4, wq, flags=re.IGNORECASE) or re.search(pattern5, wq, flags=re.IGNORECASE) or re.search(pattern6, wq) or re.search(pattern7, wq, flags=re.IGNORECASE):
#         lisst.pop(ind)
#     ind+=1
#     print(ind)
#
#     ind = ind+1



#
#
#
# browser.find_element_by_xpath("/html/body/div[1]/section/main/article/div[2]/div[1]/div/form/div/div[2]/div/label/input").send_keys("ShitHateZIS")
#
# browser.find_element_by_xpath("/html/body/div[1]/section/main/article/div[2]/div[1]/div/form/div/div[3]/button/div").click()
#
# browser.find_element_by_xpath("/html/body/div[1]/section/main/div/div/div/section/div/button").click()
#
# browser.get("https://www.instagram.com/p/CKtR_OEDB7W/")
#
#
# # browser.find_element_by_name("passwd")
# browser.find_element_by_xpath("/html/body/div[5]/div[2]/div/article/header/div[2]/div[1]/div[2]/button").click()
#
#
#
# pickle.dump( browser.get_cookies() , open("cookies.pkl","wb"))
# browser.close()
# browser FirefoxQ
# browser.get('https://yandex.ru/')
# cookies pickle.load(open("cookies.pkl", "rb"))
# for cookie	cookies:
# browser.addcookie(cookie) browser.get('https://yandex.ru/')
# fro, selenium.webdriver.firefox.options lpor Options
# opts	Options()
# opts.set_headless()
# assert opts.headless
# browser	Firefox(options-opts)
# browser.get('https://yandex.ru/')
# cookies pickle.load(open("cookies.pkl", "rb"))
# for cookie	cookies:
# browser.addcookie(cookie) browser.get( 'https://yandex.ru/')

/html/body/div[2]/section[2]/div/div[1]/div/div/a

/html/body/div[2]/header/div/div[2]/form/input[1]
/html/body/div[2]/header/div/div[2]/div/div/a



# //Get download link
# String downloadLink = driver
#          .findElement(By.cssSelector("main#content a.btn"))
#          .getAttribute("href");
#
# //Set file to save
# File fileToSave = new File("/path/to/file.zip");
#
# //Download file using default org.apache.http client
# CloseableHttpClient httpClient = HttpClients.createDefault();
# HttpGet httpGet = new HttpGet(downloadLink);
# HttpResponse response = httpClient.execute(httpGet, new BasicHttpContext());
#
# //Save file on disk
# copyInputStreamToFile(response.getEntity().getContent(), fileToSave);